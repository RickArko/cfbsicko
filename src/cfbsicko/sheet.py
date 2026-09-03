"""Read the Week 1 Google-sheet export."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from cfbsicko.parse import parse_slate


@dataclass(frozen=True)
class SheetPlayer:
    display_name: str
    picks: tuple[str, ...]


@dataclass(frozen=True)
class MasterSheet:
    slate_text: str
    players: tuple[SheetPlayer, ...]


def read_master_sheet(path: Path) -> MasterSheet:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        picks_ws = _sheet(wb, "Week 1")
        lines_ws = _sheet(wb, "WEEK 1 LINES")
        players = _read_players(picks_ws)
        slate_text = _read_slate_text(lines_ws)
        return MasterSheet(slate_text=slate_text, players=players)
    finally:
        wb.close()


def _sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    raise KeyError(f"Workbook is missing tab {name!r}; have {wb.sheetnames}")


def _read_players(ws) -> tuple[SheetPlayer, ...]:
    header = [cell.value for cell in next(ws.iter_rows(min_row=4, max_row=4))]
    names = [str(value).strip() for value in header[2:] if value]
    columns: list[list[str]] = [[] for _ in names]
    for row in ws.iter_rows(min_row=5, max_row=9, values_only=True):
        values = list(row[2 : 2 + len(names)])
        for idx, value in enumerate(values):
            columns[idx].append("" if value is None else str(value).strip())
    return tuple(SheetPlayer(display_name=name, picks=tuple(columns[idx])) for idx, name in enumerate(names))


def _read_slate_text(ws) -> str:
    lines: list[str] = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        for value in row:
            if value is None:
                continue
            text = str(value).strip()
            if text:
                lines.append(text)
                break
    parsed = parse_slate("\n".join(lines))
    if not parsed:
        raise ValueError("WEEK 1 LINES tab produced zero games")
    return "\n".join(lines)
